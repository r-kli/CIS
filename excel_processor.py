import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Color
# Standard cell formatting only
import re
from difflib import SequenceMatcher

class ExcelComparator:
    def __init__(self, file1, file2):
        self.file1 = file1
        self.file2 = file2
        self.differences_df = pd.DataFrame()
        self.green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        self.red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
        self.no_wrap_alignment = Alignment(wrap_text=False)
        self.changed_text_font = Font(bold=True, color='0000FF')  # Blue, bold font
        self.deleted_text_font = Font(strike=True)  # Strikethrough font for deleted text

    def find_text_differences(self, text1, text2):
        """Find both deletions and insertions between two text strings."""
        if pd.isna(text1) and pd.isna(text2):
            return {'deletions': [], 'insertions': []}
        text1 = str(text1) if not pd.isna(text1) else ""
        text2 = str(text2) if not pd.isna(text2) else ""
        
        matcher = SequenceMatcher(None, text1, text2)
        differences = {'deletions': [], 'insertions': []}
        
        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == 'delete':
                differences['deletions'].append((i1, i2, text1[i1:i2]))
            elif tag == 'insert':
                differences['insertions'].append((j1, j2, text2[j1:j2]))
            elif tag == 'replace':
                differences['deletions'].append((i1, i2, text1[i1:i2]))
                differences['insertions'].append((j1, j2, text2[j1:j2]))
                
        return differences

    def create_reg_key(self, row, section_col, rec_col=None):
        """Create a regulation key that handles empty/NaN recommendation values."""
        section = str(row[section_col]).strip()
        rec = ""
        if rec_col is not None and rec_col in row.index:
            rec_value = row[rec_col]
            if pd.notna(rec_value) and str(rec_value).strip():
                rec = str(rec_value).strip()
        return f"{section}_{rec}"

    def compare_files(self, progress_callback=None):
        """Compare two Excel files and identify differences."""
        xlsx1 = pd.ExcelFile(self.file1)
        xlsx2 = pd.ExcelFile(self.file2)
        
        all_differences = []
        total_sheets = len(xlsx1.sheet_names)
        
        for idx, sheet_name in enumerate(xlsx1.sheet_names):
            if sheet_name in xlsx2.sheet_names:
                df1 = pd.read_excel(xlsx1, sheet_name)
                df2 = pd.read_excel(xlsx2, sheet_name)
                
                section_col = df1.columns[0]
                rec_col = df1.columns[1] if len(df1.columns) > 1 else None
                
                df1_dict = {}
                df2_dict = {}
                
                for _, row in df1.iterrows():
                    key = self.create_reg_key(row, section_col, rec_col)
                    df1_dict[key] = row
                
                for _, row in df2.iterrows():
                    key = self.create_reg_key(row, section_col, rec_col)
                    df2_dict[key] = row
                
                for reg_key in set(df1_dict.keys()) | set(df2_dict.keys()):
                    if reg_key in df1_dict and reg_key in df2_dict:
                        row1 = df1_dict[reg_key]
                        row2 = df2_dict[reg_key]
                        
                        for col in df1.columns:
                            if row1[col] != row2[col] and not (pd.isna(row1[col]) and pd.isna(row2[col])):
                                differences = self.find_text_differences(row1[col], row2[col])
                                # Create a display-friendly version of the differences
                                diff_summary = []
                                if differences['deletions']:
                                    deleted = ', '.join(d[2] for d in differences['deletions'])
                                    diff_summary.append(f"Removed: {deleted}")
                                if differences['insertions']:
                                    inserted = ', '.join(i[2] for i in differences['insertions'])
                                    diff_summary.append(f"Added: {inserted}")
                                
                                all_differences.append({
                                    'Sheet': sheet_name,
                                    'Regulation': reg_key,
                                    'Column': col,
                                    'Old_Value': row1[col],
                                    'New_Value': row2[col],
                                    'Changes': ' | '.join(diff_summary) if diff_summary else 'Value changed',
                                    '_differences': differences  # Keep the detailed differences for Excel generation
                                })
                    else:
                        if reg_key in df1_dict:
                            row = df1_dict[reg_key]
                            all_differences.append({
                                'Sheet': sheet_name,
                                'Regulation': reg_key,
                                'Column': 'Status',
                                'Old_Value': 'Present',
                                'New_Value': 'Removed',
                                'Text_Differences': []
                            })
                        else:
                            row = df2_dict[reg_key]
                            all_differences.append({
                                'Sheet': sheet_name,
                                'Regulation': reg_key,
                                'Column': 'Status',
                                'Old_Value': 'Missing',
                                'New_Value': 'Added',
                                'Text_Differences': []
                            })
            
            if progress_callback:
                progress_callback((idx + 1) / total_sheets)
        
        # Store complete differences internally
        self.differences_df = pd.DataFrame(all_differences)
        # Return display version without internal fields
        display_df = self.differences_df.drop(columns=['_differences'])
        return display_df

    def generate_output_excel(self):
        """Generate formatted Excel output maintaining original structure with differences."""
        xlsx1 = pd.ExcelFile(self.file1)
        xlsx2 = pd.ExcelFile(self.file2)
        
        wb = Workbook()
        wb.remove(wb.active)
        
        for sheet_name in xlsx1.sheet_names:
            if sheet_name in xlsx2.sheet_names:
                df1 = pd.read_excel(xlsx1, sheet_name)
                df2 = pd.read_excel(xlsx2, sheet_name)
                
                ws = wb.create_sheet(sheet_name)
                ws.sheet_format.defaultRowHeight = 20
                
                for col_idx, col_name in enumerate(df1.columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.alignment = self.no_wrap_alignment
                    ws.row_dimensions[1].height = 20
                
                section_col = df1.columns[0]
                rec_col = df1.columns[1] if len(df1.columns) > 1 else None
                
                sheet_differences = self.differences_df[self.differences_df['Sheet'] == sheet_name]
                diff_regulations = set(sheet_differences['Regulation'].unique())
                
                output_row = 2
                
                reg_set1 = set(self.create_reg_key(row, section_col, rec_col)
                             for _, row in df1.iterrows())
                reg_set2 = set(self.create_reg_key(row, section_col, rec_col)
                             for _, row in df2.iterrows())
                
                for _, row in df1.iterrows():
                    reg_key = self.create_reg_key(row, section_col, rec_col)
                    
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=output_row, column=col_idx, value=value)
                        cell.alignment = self.no_wrap_alignment
                        if reg_key not in reg_set2:
                            cell.fill = self.red_fill
                    
                    ws.row_dimensions[output_row].height = 20
                    
                    if reg_key in diff_regulations and reg_key in reg_set2:
                        # Find the corresponding row in df2
                        new_row = next(row2 for _, row2 in df2.iterrows() 
                                     if self.create_reg_key(row2, section_col, rec_col) == reg_key)
                        diff_info = sheet_differences[
                            (sheet_differences['Regulation'] == reg_key)
                        ]
                        
                        # Write each column value, comparing with new_row for differences
                        for col_idx, (old_value, new_value) in enumerate(zip(row, new_row), 1):
                            cell = ws.cell(row=output_row, column=col_idx, value=old_value)
                            cell.alignment = self.no_wrap_alignment
                            
                            col_name = df1.columns[col_idx-1]
                            col_diffs = diff_info[diff_info['Column'] == col_name]
                            
                            if not col_diffs.empty and col_name not in [section_col, rec_col]:
                                if pd.notna(old_value) or pd.notna(new_value):
                                    cell.fill = self.green_fill
                                    text_diffs = col_diffs.iloc[0]['_differences']
                                    
                                    deletions = text_diffs['deletions']
                                    insertions = text_diffs['insertions']
                                    
                                    if deletions or insertions:
                                        # Get the full old and new text
                                        old_text = str(old_value) if pd.notna(old_value) else ""
                                        new_text = str(new_value) if pd.notna(new_value) else ""
                                        
                                        # Create text with highlighted changes
                                        if deletions or insertions:
                                            # Get the full old and new text
                                            old_text = str(old_value) if pd.notna(old_value) else ""
                                            new_text = str(new_value) if pd.notna(new_value) else ""
                                            
                                            # Process old text - highlight deleted parts in red
                                            for d_start, d_end, d_text in deletions:
                                                old_text = old_text.replace(d_text, f"*{d_text}*")
                                            
                                            # Process new text - highlight inserted parts in red
                                            for i_start, i_end, i_text in insertions:
                                                new_text = new_text.replace(i_text, f"*{i_text}*")
                                            
                                            # Combine with arrow separator
                                            cell.value = f"{old_text} --> {new_text}"
                                            
                                            # Keep original text without markers
                                            clean_old_text = old_text.replace('*', '')
                                            clean_new_text = new_text.replace('*', '')
                                            
                                            # Set cell value with arrow separator
                                            cell.value = f"{clean_old_text} --> {clean_new_text}"
                                            
                                            # Apply only cell background highlighting
                                            cell.fill = self.green_fill
                    
                    output_row += 1
                
                new_regs = reg_set2 - reg_set1
                if new_regs:
                    for reg_key in sorted(new_regs):
                        for _, new_row in df2.iterrows():
                            if self.create_reg_key(new_row, section_col, rec_col) == reg_key:
                                for col_idx, value in enumerate(new_row, 1):
                                    cell = ws.cell(row=output_row, column=col_idx, value=value)
                                    cell.alignment = self.no_wrap_alignment
                                    cell.fill = self.green_fill
                                
                                ws.row_dimensions[output_row].height = 20
                                output_row += 1
                                break
                
                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        return wb
